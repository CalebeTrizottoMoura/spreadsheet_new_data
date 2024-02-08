import openpyxl
import os
from datetime import datetime

TOTALS_ROWS_IN_SPREADSHEET = 1033

def clear_terminal():
    os.system('cls' if os.name == 'nt' else 'clear')

def row_function():
    existing_row = None
    for row in spreadsheet_new_data.iter_rows(min_row=2, max_row=spreadsheet_new_data.max_row):
        if row[0].value == date and row[1].value == time:
            existing_row = row
            return existing_row

def create_data(i, spreadsheet):
    date_time = spreadsheet[f"A{i}"].value
    date_time_formated = datetime.strptime(str(date_time), "%Y-%m-%d %H:%M:%S")
    date = date_time_formated.strftime("%d/%m/%Y")
    time = date_time_formated.strftime("%H:%M")

    variation_value = spreadsheet[f"E{i}"].value / spreadsheet[f"B{i}"].value - 1
    average_run_up = spreadsheet[f"C{i}"].value / spreadsheet[f"B{i}"].value - 1 
    drawdown = spreadsheet[f"D{i}"].value / spreadsheet[f"B{i}"].value - 1
    return date, time, variation_value, average_run_up, drawdown

def insert_value(spreadsheet_new_data, date, time, variation_value: float, average_run_up: float, drawdown: float):
    existing_row = row_function()
    if existing_row:
        existing_row[0].value = date
        existing_row[1].value = time
        existing_row[2].value = variation_value
        existing_row[3].value = average_run_up
        existing_row[4].value = drawdown
    else:
        spreadsheet_new_data.append([date, time, variation_value, average_run_up, drawdown])

def arquive():
        path = r"C:\Users\conta\OneDrive\Área de Trabalho\análise-b3\itau-1h\Novo(a) Planilha do Microsoft Excel.xlsx"
        xlsx_file = openpyxl.load_workbook(path)
        spreadsheet = xlsx_file.active

        new_data_name = "New Data"

        if new_data_name not in xlsx_file.sheetnames:
            xlsx_file.create_sheet(new_data_name)
            spreadsheet_new_data = xlsx_file[new_data_name]
            spreadsheet_new_data.append(["DATA", "HORA", "VARIAÇÃO", "MÁXIMO GANHO", "MÁXIMA PERDA"])
        else:
            spreadsheet_new_data = xlsx_file[new_data_name]
        return path, xlsx_file, spreadsheet, spreadsheet_new_data

def count_negative_positive(count_postive, count_negative, total):
    variation_new_data = spreadsheet_new_data[f"C{i}"].value
    if variation_new_data is not None:
        total += 1
        variation_new_data = float(variation_new_data)
        if variation_new_data >= 0:
            count_postive += 1
        else: 
            count_negative += 1
    return count_postive, count_negative, total

count_postive = 0
count_negative = 0
total = 0

path, xlsx_file, spreadsheet, spreadsheet_new_data = arquive()

for i in range(2, TOTALS_ROWS_IN_SPREADSHEET):
    date, time, variation_value, average_run_up, drawdown = create_data(i, spreadsheet)
    insert_value(spreadsheet_new_data, date, time, variation_value, average_run_up, drawdown)
    count_postive, count_negative, total = count_negative_positive(count_postive, count_negative, total)
    
xlsx_file.save(path)

clear_terminal()
print(f"Total: {total}")
print(f"Postivo: {count_postive}")
print(f"Negativo: {count_negative}")
